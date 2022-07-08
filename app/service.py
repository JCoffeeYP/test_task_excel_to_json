import json
import logging
import os
import shutil
from datetime import datetime, time
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__file__)


def save_upload_file(upload_file: UploadFile, destination: Path) -> None:
    try:
        with destination.open("wb") as buffer:
            shutil.copyfileobj(upload_file.file, buffer)
            logger.info(f"Файл {upload_file.filename} успешно сохранён")
    finally:
        upload_file.file.close()


def data_type_definition(value: Any):

    if isinstance(value, datetime) or isinstance(value, time):
        logger.info(f"{value} формат даты")
        return value.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
    elif isinstance(value, str):
        try:
            formatted_date = datetime.strptime(value, "%Y-%m-%d %H:%M:%S.%f")
            return formatted_date.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"
        except ValueError:
            logger.info(f"{value} неверный формат даты")
            pass
        try:
            formatted_value = json.loads(value)
            for key, obj in formatted_value.items():
                formatted_value[key] = data_type_definition(obj)
            return formatted_value
        except (TypeError, AttributeError, json.decoder.JSONDecodeError):
            logger.info(f"{value} неудачная попытка преобразовать в json")
            pass
    elif isinstance(value, dict):
        logger.info(f"{value} формат словаря")
        for key, obj in value.items():
            value[key] = data_type_definition(obj)
    return value


def check_empty_row(row: dict) -> bool:
    """
    Данная проверка нужна для того, чтобы не учитывать пустые строки в конце файла
    """
    empty_cell_count = 0
    for value in row.values():
        if not value:
            empty_cell_count += 1
    if empty_cell_count == len(row):
        return True
    return False


def parse_excel_file(filepath: Path) -> dict:
    wb = load_workbook(filename=filepath)

    result: dict = {}

    for sheet in wb:
        result[sheet.title]: list = []
        wb.active = sheet
        ws: Worksheet = wb.active
        last_column: int = len(list(ws.columns))
        last_row: int = len(list(ws.rows))

        for row in range(1, last_row + 1):
            row_value: dict = {}
            for column in range(1, last_column + 1):
                column_letter: str = get_column_letter(column)
                intermediate_value: Any = ws[f"{column_letter}{row}"].value
                if row > 1 and ws[f"{column_letter}1"].value is not None:
                    row_value[ws[f"{column_letter}1"].value] = data_type_definition(
                        intermediate_value
                    )
            if row_value and not check_empty_row(row_value):
                result[sheet.title].append(row_value)
    wb.close()
    return result


def convert_excel_file_to_json(tmp_filepath: Path) -> None:
    parsed_excel: dict = parse_excel_file(tmp_filepath)
    data: str = json.dumps(parsed_excel, indent=4, ensure_ascii=False)
    with open(f"{tmp_filepath.with_suffix('.json')}", "w", encoding="utf-8") as f:
        f.write(data)


def check_existing_files(upload_file: UploadFile):
    storage_dir = os.environ.get("STORAGE_DIR", "json_output_data")
    current_directory = Path("..", storage_dir)
    json_filename: Path = current_directory.joinpath(
        Path(upload_file.filename).with_suffix(".json")
    )
    default_filename: Path = current_directory.joinpath(Path(upload_file.filename))

    for file in current_directory.iterdir():
        if file == json_filename or file == default_filename:
            return False
    return True


def validate_input_file_format(upload_file: UploadFile):
    supported_formats = [".xlsx", ".xlsm", ".xltx", ".xltm"]
    if Path(upload_file.filename).suffix not in supported_formats:
        return False
    return True
